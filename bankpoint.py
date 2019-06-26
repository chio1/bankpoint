import discord
import asyncio
import openpyxl
import os

client = discord.Client()

@client.event
async def on_ready():
    print('온')
    print('------')
    game = discord.Game("!명령어 | By Chio")
    await client.change_presence(status=discord.Status.online , activity=game)



@client.event
async def on_message(message):
    owner_carpe = "333910993216208896"
    owner_chio = "267558400861143040"
    owner_anjihyeon = "364336734323867648"
    if message.content.startswith("!포인트주기"):
        if str(message.author.id) == owner_anjihyeon:
            if str(message.content[7:]) == "":
                await message.channel.send(embed =discord.Embed(title="포인트주기", description="<명령어> !포인트주기 (고유번호)" , color=0x00ff00))
            else:

                file = openpyxl.load_workbook("포인트.xlsx")
                sheet = file.active

                while True:
                    if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                        sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) + 1
                        file.save("포인트.xlsx")
                        await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))


                        break
                    if sheet["A" + str(message.content[7:])].value == None:
                        sheet["A" + str(message.content[7:])].value = str(message.content[7:])
                        sheet["B" + str(message.content[7:])].value = 1
                        file.save("포인트.xlsx")
                        await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))
                        break
        elif str(message.author.id) == owner_chio :
                if str(message.content[7:]) == "":
                    await message.channel.send(embed =discord.Embed(title="포인트주기", description="<명령어> !포인트주기 (고유번호)" , color=0x00ff00))
                else:

                    file = openpyxl.load_workbook("포인트.xlsx")
                    sheet = file.active

                    while True:
                        if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                            sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) + 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))
                            break
                        if sheet["A" + str(message.content[7:])].value == None:
                            sheet["A" + str(message.content[7:])].value = str(message.content[7:])
                            sheet["B" + str(message.content[7:])].value = 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))
                            break
        elif str(message.author.id) == owner_carpe:
                if str(message.content[7:]) == "":
                    await message.channel.send(embed =discord.Embed(title="포인트주기", description="<명령어> !포인트주기 (고유번호)" , color=0x00ff00))
                else:

                    file = openpyxl.load_workbook("포인트.xlsx")
                    sheet = file.active

                    while True:
                        if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                            sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) + 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))


                            break
                        if sheet["A" + str(message.content[7:])].value == None:
                            sheet["A" + str(message.content[7:])].value = str(message.content[7:])
                            sheet["B" + str(message.content[7:])].value = 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트주기", description="축하드립니다 진급포인트 1획득 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))
                            break
        else:
            await message.channel.send(embed=discord.Embed(title="포인트주기", description="사용불가", color=0x00ff00))
    if message.content.startswith("!포인트확인"):
        if str(message.content[7:]) == "":
            await message.channel.send(embed =discord.Embed(title="포인트확인", description="<명령어> !포인트확인 (고유번호)",color=0x00ff00))
        else:

            file = openpyxl.load_workbook("포인트.xlsx")
            sheet = file.active
            while True:
                if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                    await message.channel.send(embed =discord.Embed(title="포인트확인", description=str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))
                    break
                else:
                    await message.channel.send(embed =discord.Embed(title="포인트확인", description=str(message.content[7:]) + "님의 누적 포인트 : 0 ",color=0x00ff00) )
                    break
    if message.content.startswith("!포인트뺏기"):
        if str(message.author.id) ==  owner_carpe:
            if str(message.content[7:]) == "":
                await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="<명령어> !포인트뺏기 (고유번호)" , color=0x00ff00))
            else:

                file = openpyxl.load_workbook("포인트.xlsx")
                sheet = file.active

                while True:
                    if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                        sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) - 1
                        file.save("포인트.xlsx")
                        await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="분발해세요 진급포인트 1뺏김 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))


                        break
                    if sheet["A" + str(message.content[7:])].value == None:
                        await message.channel.send(embed=discord.Embed(title="포인트뺏기", description="찾을 수 없습니다.", color=0x00ff00))
                        break
        elif str(message.author.id) == owner_chio:
                if str(message.content[7:]) == "":
                    await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="<명령어> !포인트뺏기 (고유번호)" , color=0x00ff00))
                else:

                    file = openpyxl.load_workbook("포인트.xlsx")
                    sheet = file.active

                    while True:
                        if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                            sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) - 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="분발해세요 진급포인트 1뺏김 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))


                            break
                        if sheet["A" + str(message.content[7:])].value == None:
                            await message.channel.send(embed=discord.Embed(title="포인트뺏기", description="찾을 수 없습니다.", color=0x00ff00))
                            break

        elif str(message.author.id) == owner_anjihyeon:
                if str(message.content[7:]) == "":
                    await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="<명령어> !포인트뺏기 (고유번호)" , color=0x00ff00))
                else:

                    file = openpyxl.load_workbook("포인트.xlsx")
                    sheet = file.active

                    while True:
                        if sheet["A" + str(message.content[7:])].value == str(message.content[7:]):
                            sheet["B" + str(message.content[7:])].value = int(sheet["B" + str(message.content[7:])].value) - 1
                            file.save("포인트.xlsx")
                            await message.channel.send(embed =discord.Embed(title="포인트뺏기", description="분발해세요 진급포인트 1뺏김 \n " + str(message.content[7:]) +"님의 누적 포인트 : " + str(sheet["B" + str(message.content[7:])].value),color=0x00ff00))


                            break
                        if sheet["A" + str(message.content[7:])].value == None:
                            await message.channel.send(embed=discord.Embed(title="포인트뺏기", description="찾을 수 없습니다.", color=0x00ff00))

        else:
            await message.channel.send(embed=discord.Embed(title="포인트뺏기", description="사용불가", color=0x00ff00))
    if message.content.startswith("!명령어"):
        await message.channel.send(embed=discord.Embed(title="명령어", description= "\n***<직원>***\n ```!포인트확인``` \n\n\n***<이사회>(직원사용불가)***\n```!포인트주기\n!포인트뺏기```", color=0x00ff00))

client.run('NTkyNjc1NjA0MzExMDQ4MjAy.XRNBNw.CEWTffp3RZY6Z6lHaebINSx-29s')
access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
